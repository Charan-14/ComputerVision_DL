import cv2 as cv
import pytesseract
import openpyxl

wb = openpyxl.Workbook()
sheet = wb.active

#CV based Text Detection
img = cv.imread("alphanumeric.png")

gray = cv.cvtColor(img, cv.COLOR_BGR2GRAY)

# Performing OTSU threshold 
ret, thresh1 = cv.threshold(gray, 75, 255, cv.THRESH_BINARY_INV) 
  
# Specify structure shape and kernel
rect_kernel = cv.getStructuringElement(cv.MORPH_RECT, (11, 11)) 
  
# Appplying dilation on the threshold image 
dilation = cv.dilate(thresh1, rect_kernel, iterations = 1) 
  
# Finding contours 
contours, hierarchy = cv.findContours(dilation, cv.RETR_EXTERNAL, cv.CHAIN_APPROX_NONE) 
  
# Creating a copy of image 
im2 = img.copy() 
textleft = []
textright = []
for cnt in contours: 
    x, y, w, h = cv.boundingRect(cnt) 
      
    # Drawing a rectangle on copied image 
    rect = cv.rectangle(im2, (x, y), (x + w, y + h), (0, 255, 0), 2) 
      
    # Cropping the text block for giving input to OCR 
    cropped = im2[y:y + h, x:x + w] 
      
    # Apply OCR on the cropped image 
    string = pytesseract.image_to_string(cropped)
    if len(string)==0:
        continue
    else:
        
        splitted = string.split("\n")
        if x<im2.shape[0]/2:
            textleft.append(splitted)
        else:
            textright.append(splitted)
cv.imshow("conts", im2)
cv.imshow("tresh", thresh1)
print(textright)
print(textleft)
cv.waitKey(0)
cv.destroyAllWindows()

for i,text in enumerate(textleft):
    c1 = sheet.cell(row = 1, column = 1) 
    c1.value = "Left Coloumn"
    x = sheet.cell(row = i+1, column = 1)
    if len(text)>1:
        x.value = text[0] + text[1]
    else:
        x.value = text[0]
        
for i,text in enumerate(textright):
    c1 = sheet.cell(row = 1, column = 7) 
    c1.value = "Left Coloumn"
    x = sheet.cell(row = i+1, column = 7)
    if len(text)>1:
        x.value = text[0] + text[1]
    else:
        x.value = text[0]
wb.save("/home/blackpanther/Desktop/Projects/Text Detection and Recognition/extracted_text.xlsx")

#DL Based Text Detection

#img =  cv.resize(img, (320,320))

#inpWidth = img.shape[0]
#inpHeight = img.shape[1]
# #Model create from weights
# net = cv.dnn.readNet("frozen_east_text_detection.pb")

# #Create 4D blob from image
# blob = cv.dnn.blobFromImage(img, 1, (inpWidth, inpHeight), (123.68, 116.78, 103.94), True, False)

# #Extracting scores and bounding box dimensions
# outputLayers = []

# outputLayers.append("feature_fusion/Conv_7/Sigmoid")
# outputLayers.append("feature_fusion/concat_3")

# #Input to network
# net.setInput(blob)

# #Output of network
# output = net.Forward(outputLayers)

# scores = output[0]
# geometry = output[1]

# #NMS
# [boxes, confidences] = decode(scores, geometry, confThreshold)

# indices = cv.dnn.NMSBoxesRotated(boxes, confidences, confThreshold, nmsThreshold)

# print(indices)

